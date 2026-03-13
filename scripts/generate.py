#!/usr/bin/env python3
"""
Скрипт объединения товарных отчётов Энтерсайт в один файл с подённой разбивкой.
Подробная логика описана в SKILL.md рядом со скриптом.

Использование:
    python3 generate.py [--data-dir PATH] [--output PATH]

По умолчанию:
    --data-dir = <корень проекта>/данные Энтерсайт/
    --output   = <data-dir>/Итоговые_данные_Энтерсайт.xlsx
"""

import argparse
import io
import os
import re
import sys
from datetime import datetime, timedelta

import openpyxl
from openpyxl import Workbook


def parse_period_from_filename(filename: str):
    """Извлекает даты начала и конца периода из имени файла."""
    pattern = r"с_(\d{2})_(\d{2})_(\d{4})_по_(\d{2})_(\d{2})_(\d{4})"
    match = re.search(pattern, filename)
    if not match:
        return None, None
    d1, m1, y1, d2, m2, y2 = match.groups()
    start_date = datetime(int(y1), int(m1), int(d1))
    end_date = datetime(int(y2), int(m2), int(d2))
    return start_date, end_date


def read_file_data(filepath_or_stream):
    """Читает данные товаров из xlsx (путь или file-like).
    Пропускает строки с пустым артикулом или нулевым количеством реализации.
    Колонки: B=Артикул, C=Товар, L=Реализация:Количество, M=Реализация:Закуп, N=Реализация:Розница
    """
    wb = openpyxl.load_workbook(filepath_or_stream, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        article = row[1]
        name = row[2]
        quantity = row[11]
        cost_sum = row[12]
        revenue = row[13]

        if not article or article == "" or quantity is None or quantity == 0:
            continue

        try:
            quantity = float(quantity)
            cost_sum = float(cost_sum) if cost_sum else 0.0
            revenue = float(revenue) if revenue else 0.0
        except (ValueError, TypeError):
            continue

        if quantity == 0:
            continue

        rows.append({
            "article": str(article).strip(),
            "name": str(name).strip() if name else "",
            "quantity": quantity,
            "cost_sum": cost_sum,
            "revenue": revenue,
        })
    wb.close()
    return rows


def merge_entersite_from_uploads(files):
    """
    Объединяет загруженные xlsx Энтерсайт в один файл.
    files: [(filename, bytes), ...]
    Возвращает (excel_bytes, info_message) или (None, error_message).
    """
    file_records = []
    for fname, content in files:
        if not fname.endswith(".xlsx") or fname.startswith("~$") or fname.startswith("Итог"):
            continue
        start_date, end_date = parse_period_from_filename(fname)
        if start_date is None:
            continue
        file_records.append({
            "filename": fname,
            "content": content,
            "start_date": start_date,
            "end_date": end_date,
        })

    file_records.sort(key=lambda x: x["start_date"])
    if not file_records:
        return None, "Нет подходящих файлов (имя должно содержать период: с_ДД_ММ_ГГГГ_по_ДД_ММ_ГГГГ)"

    cost_data = {}
    sales_data = {}

    for fr in file_records:
        try:
            stream = io.BytesIO(fr["content"])
            data = read_file_data(stream)
        except Exception as e:
            return None, f"Ошибка чтения {fr['filename']}: {e}"

        start = fr["start_date"]
        end = fr["end_date"]
        num_days = (end - start).days + 1

        for item in data:
            article = item["article"]
            name = item["name"]
            cost_per_unit = item["cost_sum"] / item["quantity"]
            qty_per_day = item["quantity"] / num_days
            price_per_unit = item["revenue"] / item["quantity"]

            if article not in cost_data:
                cost_data[article] = {}
            if article not in sales_data:
                sales_data[article] = {}

            for day_offset in range(num_days):
                current_date = start + timedelta(days=day_offset)
                cost_data[article][current_date] = {"cost": round(cost_per_unit, 2), "name": name}
                sales_data[article][current_date] = {
                    "quantity": round(qty_per_day, 6),
                    "price": round(price_per_unit, 2),
                    "name": name,
                }

    wb = Workbook()
    ws_cost = wb.active
    ws_cost.title = "Себестоимость"
    ws_cost.append(["product_id", "stock", "cost", "date"])
    for article in sorted(cost_data.keys()):
        for date in sorted(cost_data[article].keys()):
            entry = cost_data[article][date]
            ws_cost.append([article, 1, entry["cost"], date.strftime("%Y-%m-%d")])

    ws_sales = wb.create_sheet("Продажи")
    ws_sales.append(["product_id", "name_full", "quantity", "recorded_on", "price"])
    for article in sorted(sales_data.keys()):
        for date in sorted(sales_data[article].keys()):
            entry = sales_data[article][date]
            ws_sales.append([
                article, entry["name"], entry["quantity"],
                date.strftime("%Y-%m-%d"), entry["price"],
            ])

    buf = io.BytesIO()
    wb.save(buf)
    cost_rows = sum(len(cost_data[a]) for a in cost_data)
    sales_rows = sum(len(sales_data[a]) for a in sales_data)
    info = f"Обработано файлов: {len(file_records)}. Себестоимость: {cost_rows} строк, Продажи: {sales_rows} строк."
    return buf.getvalue(), info


def main():
    # Определяем корневую папку проекта (на 2 уровня выше scripts/)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    skill_dir = os.path.dirname(script_dir)

    # Пытаемся найти корень проекта — ищем папку "данные Энтерсайт"
    # Сначала проверяем переданные аргументы, затем CWD, затем типичные пути
    parser = argparse.ArgumentParser(description="Объединение товарных отчётов Энтерсайт")
    parser.add_argument("--data-dir", type=str, default=None,
                        help="Путь к папке с исходными xlsx файлами")
    parser.add_argument("--output", type=str, default=None,
                        help="Путь к итоговому xlsx файлу")
    args = parser.parse_args()

    # Определяем data_dir
    if args.data_dir:
        data_dir = os.path.abspath(args.data_dir)
    else:
        # Проверяем CWD
        cwd = os.getcwd()
        candidate = os.path.join(cwd, "данные Энтерсайт")
        if os.path.isdir(candidate):
            data_dir = candidate
        else:
            print("Ошибка: не удалось найти папку 'данные Энтерсайт'.")
            print("Укажите путь через --data-dir")
            sys.exit(1)

    if not os.path.isdir(data_dir):
        print(f"Ошибка: папка не существует: {data_dir}")
        sys.exit(1)

    output_file = args.output or os.path.join(data_dir, "Итоговые_данные_Энтерсайт.xlsx")

    # 1. Собираем файлы с периодами
    file_records = []
    for fname in os.listdir(data_dir):
        if not fname.endswith(".xlsx") or fname.startswith("~$") or fname.startswith("Итог"):
            continue
        start_date, end_date = parse_period_from_filename(fname)
        if start_date is None:
            print(f"  Пропуск (нет периода в имени): {fname}")
            continue
        file_records.append({
            "filename": fname,
            "filepath": os.path.join(data_dir, fname),
            "start_date": start_date,
            "end_date": end_date,
        })

    file_records.sort(key=lambda x: x["start_date"])

    print(f"Найдено файлов: {len(file_records)}")
    for fr in file_records:
        print(f"  {fr['start_date'].strftime('%d.%m.%Y')} - "
              f"{fr['end_date'].strftime('%d.%m.%Y')}: {fr['filename']}")

    if not file_records:
        print("Нет файлов для обработки.")
        sys.exit(1)

    # 2. Обработка: строим дневные данные с приоритетом более позднего файла
    cost_data = {}   # {article: {date: {"cost": ..., "name": ...}}}
    sales_data = {}  # {article: {date: {"quantity": ..., "price": ..., "name": ...}}}

    for fr in file_records:
        print(f"\nОбработка: {fr['filename']}")
        data = read_file_data(fr["filepath"])
        print(f"  Товаров с продажами: {len(data)}")

        start = fr["start_date"]
        end = fr["end_date"]
        num_days = (end - start).days + 1

        for item in data:
            article = item["article"]
            name = item["name"]
            cost_per_unit = item["cost_sum"] / item["quantity"]
            qty_per_day = item["quantity"] / num_days
            price_per_unit = item["revenue"] / item["quantity"]

            if article not in cost_data:
                cost_data[article] = {}
            if article not in sales_data:
                sales_data[article] = {}

            for day_offset in range(num_days):
                current_date = start + timedelta(days=day_offset)
                cost_data[article][current_date] = {
                    "cost": round(cost_per_unit, 2),
                    "name": name,
                }
                sales_data[article][current_date] = {
                    "quantity": round(qty_per_day, 6),
                    "price": round(price_per_unit, 2),
                    "name": name,
                }

    # 3. Формируем итоговый Excel
    wb = Workbook()

    ws_cost = wb.active
    ws_cost.title = "Себестоимость"
    ws_cost.append(["product_id", "stock", "cost", "date"])

    cost_row_count = 0
    for article in sorted(cost_data.keys()):
        for date in sorted(cost_data[article].keys()):
            entry = cost_data[article][date]
            ws_cost.append([article, 1, entry["cost"], date.strftime("%Y-%m-%d")])
            cost_row_count += 1

    ws_sales = wb.create_sheet("Продажи")
    ws_sales.append(["product_id", "name_full", "quantity", "recorded_on", "price"])

    sales_row_count = 0
    for article in sorted(sales_data.keys()):
        for date in sorted(sales_data[article].keys()):
            entry = sales_data[article][date]
            ws_sales.append([
                article, entry["name"], entry["quantity"],
                date.strftime("%Y-%m-%d"), entry["price"],
            ])
            sales_row_count += 1

    print(f"\nСебестоимость: {cost_row_count} строк")
    print(f"Продажи: {sales_row_count} строк")

    wb.save(output_file)
    print(f"\nФайл сохранён: {output_file}")


if __name__ == "__main__":
    main()
