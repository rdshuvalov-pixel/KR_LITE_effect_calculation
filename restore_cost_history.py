"""
Скрипт для восстановления исторических данных себестоимости.
Для каждого товара берется самая ранняя известная дата, и генерируются строки
с датами от 22.09.2025 до этой даты с теми же значениями остатка и себестоимости.
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os

# Константы - находим файл с учетом неразрывного пробела
def find_input_file():
    """Находит файл с new_cost в имени, учитывая неразрывные пробелы."""
    for file in os.listdir('.'):
        if 'new_cost' in file and file.endswith('.xlsx'):
            return file
    raise FileNotFoundError("Не найден файл с 'new_cost' в имени")

INPUT_FILE = find_input_file()
SHEET_NAME = 'Себестоимость'
NEW_SHEET_NAME = 'Восстановленные данные'
START_DATE = pd.to_datetime('2025-09-22')


def restore_cost_history():
    """
    Восстанавливает исторические данные себестоимости для товаров.
    """
    print(f"Чтение данных из файла {INPUT_FILE}, вкладка '{SHEET_NAME}'...")
    
    # Читаем исходные данные
    df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
    
    # Преобразуем дату в datetime
    df['date'] = pd.to_datetime(df['date'])
    
    print(f"Загружено {len(df)} строк данных")
    print(f"Уникальных товаров: {df['product_id'].nunique()}")
    
    # Группируем по product_id и находим минимальную дату для каждого товара
    print("\nОбработка данных по товарам...")
    earliest_dates = df.groupby('product_id')['date'].min()
    
    # Получаем значения stock и cost на минимальную дату для каждого товара
    restored_rows = []
    
    for product_id, min_date in earliest_dates.items():
        # Получаем строку с минимальной датой для этого товара
        product_data = df[(df['product_id'] == product_id) & (df['date'] == min_date)].iloc[0]
        stock = product_data['stock']
        cost = product_data['cost']
        
        # Если минимальная дата больше 22.09.2025, генерируем исторические данные
        if min_date > START_DATE:
            # Генерируем даты от START_DATE до (min_date - 1 день)
            date_range = pd.date_range(start=START_DATE, end=min_date - timedelta(days=1), freq='D')
            
            # Создаем строки для каждой даты
            for date in date_range:
                restored_rows.append({
                    'product_id': product_id,
                    'stock': stock,
                    'cost': cost,
                    'date': date
                })
    
    # Создаем DataFrame из восстановленных данных
    restored_df = pd.DataFrame(restored_rows)
    
    print(f"\nСгенерировано {len(restored_df)} строк восстановленных данных")
    print(f"Уникальных товаров с восстановленными данными: {restored_df['product_id'].nunique()}")
    
    if len(restored_df) == 0:
        print("Нет данных для восстановления (все товары имеют даты <= 22.09.2025)")
        return
    
    # Сортируем по product_id и date для удобства
    restored_df = restored_df.sort_values(['product_id', 'date']).reset_index(drop=True)
    
    # Открываем Excel файл для записи
    print(f"\nСохранение данных в новую вкладку '{NEW_SHEET_NAME}'...")
    
    # Загружаем существующий файл
    book = load_workbook(INPUT_FILE)
    
    # Удаляем вкладку, если она уже существует
    if NEW_SHEET_NAME in book.sheetnames:
        book.remove(book[NEW_SHEET_NAME])
    
    # Создаем новую вкладку
    with pd.ExcelWriter(INPUT_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # Записываем данные в новую вкладку
        restored_df.to_excel(writer, sheet_name=NEW_SHEET_NAME, index=False)
    
    print(f"Данные успешно сохранены в вкладку '{NEW_SHEET_NAME}'")
    print(f"\nСтатистика:")
    print(f"  - Всего строк: {len(restored_df)}")
    print(f"  - Уникальных товаров: {restored_df['product_id'].nunique()}")
    print(f"  - Диапазон дат: {restored_df['date'].min().date()} - {restored_df['date'].max().date()}")


if __name__ == '__main__':
    restore_cost_history()
